___author__ = 'soumukhe'
# !/usr/bin/env python
#
"""
Documentation:
#
Steps: 
#
1) Make sure you have the required python packages that are imported in the import statements 
2) make a folder/directory and put the script ( the .py file) in that folder
3) Log in to ACI, APIC controller as root.
4) cd to /data/nginx/html/doc/html
5) do a "ls | grep FAULT" ( or ls -lag | grep FAULT) and save the results  in that directory to a file called "aci-faults-lists.txt"
This file is basically a list of all the Faults for that release. The  content of the file should look like below:
#
The output of the file should look like show below:
#
FAULT-F0020.html
FAULT-F0021.html
FAULT-F0022.html
FAULT-F0023.html
<snip>
#
Or:
#
-rw-rw-r--. 1 1001 2001   10364 May 25 07:20 FAULT-F0020.html
-rw-rw-r--. 1 1001 2001    4701 May 25 07:20 FAULT-F0021.html
-rw-rw-r--. 1 1001 2001    4655 May 25 07:20 FAULT-F0022.html
-rw-rw-r--. 1 1001 2001    4770 May 25 07:20 FAULT-F0023.html
-rw-rw-r--. 1 1001 2001    3605 May 25 07:20 FAULT-F0048.html
<snip>
#
6) Run the script.  It will take approximately 3 days to compete the script.  If the script exits due to IO error or connectivity loss
re-run the script.  The script will figure out where the script left off and then contiure from there.  The spreadsheet will be created in 
the same directory.  The name of it will be "ACI-Faults.xlsx"
#
What this script does:
#----------------------
This script creates a list that stores the entire 34K file names and makes a list out of them.  
Then it iterates through the list and for each entry using Urllib to read the indivual html entries.  It then 
uses BeautifulSoup4 to format the output and save to a file.  Then the script uses regex and some other methods to get the relevant information.  
After that it uses openpyxl to save to a spreadsheet .  
# 
"""
#
#
import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import bs4 as bs
import re
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
from openpyxl.utils import get_column_letter, column_index_from_string  # import column_letter and index from string
from openpyxl.styles import Alignment
import time
from glob import glob
import json
import sys   #  only using while testing code
#
#
#
def MakeSpreadsheet():
    # open workbook and create consumer-provider  sheet, Create Header Row
    wb = openpyxl.Workbook()  # open new workbook
    wb.create_sheet(index=0, title='ACI Fault Codes')  # insert a new sheet in front
    FirstSheet = wb.get_sheet_by_name('ACI Fault Codes')  # choose this sheet
    # define the header row with the name of the fields
    FirstSheet["A1"] = "F-Code"
    FirstSheet["B1"] = "Message"
    FirstSheet["C1"] = "Explanation"
    FirstSheet["D1"] = "Severity"
    FirstSheet["E1"] = "Weight"
    FirstSheet["F1"] = "Fault Name"
    FirstSheet["G1"] = "Cause"
    FirstSheet["H1"] = "Type"
    FirstSheet["I1"] = "Raised on MO"
    FirstSheet["J1"] = "Recommended Action"
    #
    # Make them bold, and turn AutoFilter on
    #
    for i in range(1, 11):  #
        columnnum = get_column_letter(i)
        colrow = columnnum + str(1)
        cell = FirstSheet[colrow]
        cell.font = Font(size=12, name="Arial", color=colors.BLUE, italic=True, bold=True)
        FirstSheet.column_dimensions[columnnum].width = 20  # Set size of Columns to 20
        # rd = FirstSheet.row_dimensions[2]  # get dimension for row 2
        # rd.height = 100  # value in points, there is no "auto"
        if i == 2:
            FirstSheet.column_dimensions[columnnum].width = 40  # Set size of last Columns to 40
        if i == 10:
            FirstSheet.column_dimensions[columnnum].width = 160  # Set size of last Columns to 160
    #
    # Turn on Freeze Panes and Auto Filter
    FirstSheet.freeze_panes = "A2"  # Freeze Panes
    FirstSheet.auto_filter.ref = "A1:J1"  # Turn On Auto Filter
    #
    wb.save("ACI-Faults.xlsx")  # Save the excel spreadsheet
#
#
#
def get_apic_creds():
#
    while True:
        ip = raw_input("# Enter APIC IP Address e.g.10.10.10.40 : ")
        if len(ip) < 1:  ip = "10.10.10.40"
        print ("You entered"  ,   ip )
        is_OK = raw_input("# Enter \"YES\" if above is O.K: ")
        if is_OK == "YES":
          break
    #
    while True:
        username = raw_input("# Enter APIC Username e.g. apic:fallback\\admin  : ")
        if len(username) < 1:  username = "apic:fallback\\admin"
        print ("You entered"  ,   username )
        is_OK = raw_input("# Enter \"YES\" if above is O.K: ")
        if is_OK == "YES":
          break
    #
    while True:
        passwd = raw_input("# Enter APIC Password e.g. mySecretPassword101 : ")
        if len(passwd) < 1:  passwd = "mySecretPassword101"
        print ("You entered"  ,   passwd )
        is_OK = raw_input("# Enter \"YES\" if above is O.K: ")
        if is_OK == "YES":
          break
    #
    #
    creds = [ip, username, passwd]
    return creds
#
#
#
def make_list(highest_row):
    ##****************** Get Faults *************
    global fcodes
    fcodes = []
    fh = open("aci-faults-list.txt", 'r')  #  open the file for reading ( this file was the one you created in the directory with the ls command)
    #
    rowstart = 1
    for line1 in fh:
        if rowstart > (
            highest_row - 1):  # in the aci-faults-list.txt, this is where you need to start back from to make the list fcodes
            atpos = line1.find("FAULT")
            stpos = line1.find("html")
            fault_file = line1[atpos: stpos + 4]
            fcodes.append(fault_file)
            # print (rowstart)
        rowstart += 1
    # fcodes = ["FAULT-F0020.html", "FAULT-F0474.html", "FAULT-F0467.html"]  # used for testing 3 iterations only
#
#
#
if __name__ == "__main__":
#
    mydict = {}   #  define the dictionary where we will store the values obtained temporarily
    sp_exists = 0  #   For first run the spreadsheet will not exist.  We don't want to overwrite the spreadsheet, if you run the script again, in case you get disconnected or have i/o error
    # print glob("*")
    files = glob("*")   #   This will create a list called files in the current directory with the names of the files as elements
    #
    for file in files:
        if file == "ACI-Faults.xlsx":
            sp_exists = 1   #  if the spreadsheet exists then value is 1
            # print ("sp_exists", sp_exists)
    #
    #
    if sp_exists == 0 :
        MakeSpreadsheet()  # make a new spreadsheet only for 1st run, or if you delete the spreadsheet from the directory
    #
    wb = openpyxl.load_workbook('ACI-Faults.xlsx')  # load the worksheet
    FirstSheet = wb.get_sheet_by_name('ACI Fault Codes')  # choose this sheet
    #
    #  First find out row in spreadsheet where you need to start writing, you don't want to overwrite rows for subsequent runs
    highest_row = 1
    for key in FirstSheet._cells:
        if key[0] > highest_row:
            highest_row = key[0]
    #
    print ("Number of Rows already in spreadsheet: ", highest_row)
    #
    #
    row = highest_row + 1  # for continued writing to spreadsheet,  use  + 1 , because 1st row is header
    #
    #
    apic_creds = get_apic_creds()
    # print (apic_creds)
    apic_ip = apic_creds[0]
    username = apic_creds[1]
    passwd  = apic_creds[2]
    #
    #
    # sys.exit("Stopped Code Below")  # for testing only
    make_list(highest_row)   #  on subsequent runs the list should only contain the newer elements
    print (fcodes[0])  #  so you can see the Fault Code being worked on
    #
    #
    for fcode in fcodes:
        #
        # **************  use urllib3 with cookies **************
        #  if we used http instead of https, this would be much simpler, but most APICs have http disabled, so better to use https
        #
        # Json Login Credentials,  using json to obtain the cookie
        name_pwd = {"aaaUser": {"attributes": {"name": username, "pwd": passwd}}}
        json_credentials = json.dumps(name_pwd, indent=4)  # make in pretty print json format
        #
        # *************Login to APIC************************
        base_url = "https://" + apic_ip + "/api/"   #  This is the base_url for json query
        login_url = base_url + 'aaaLogin.json'      #  using the aaaLogin.json with base_url
        #
        post_response = requests.post(login_url, data=json_credentials,
                                      verify=False)  # verify = False is needed because cert is self signed for most APICS
        #
        # ***********Get Token from Login Response****************
        auth = json.loads(post_response.text)  # loads the response to a json file (dictionary), this has the token deep down in it
        # print json.dumps(auth, indent = 4)  # json dumps with indent, puts it in a dictionary of json items, look at this to figure out how to extract the token
        login_attributes = auth['imdata'][0]['aaaLogin']['attributes']  #  drill down and get the "attributes" dictionary  which is inside List "imdata" and  Dictionary "aaaLogin"
        # print (login_attributes)  # This will now print the attributes dictionary
        auth_token = login_attributes['token']  # get the value of token from the key attributes
        #
        #   Using the Extracted Token make the Cookies (directory)  with Key/Value pair  of APIC-Cookie/Token
        cookies = {}  # create empty dictionary
        cookies['APIC-Cookie'] = auth_token  # "APIC-Cookie" key is the required key for APIC, so make a dictionary with key = APIC-Cookie and value = token
        # print (cookies)  # in case you wanted to see what the cookies dictionary looks like
        # this is what it looks like:
        # 'APIC-Cookie': u'FOAjAAAAAAAAAAAAAAAAAL3I18Llae41CQE9a/Gpl+Klrz04OlZKHYa4F2g6khgyTprY3hgJI3YEmBM0kvRQtY36ur0P1+ziFYYbuxKPrGp4ctAon6V9AsfVOgHHopQiVd9YQT4qRlseFCNieW1ZQqZFO87a0RUYp8a5XdRsyw3joD4PGoL5xFcICwSwXGmRjceNhm0F+Z2RMzoKJd/wfA=='}
        #
        # now use the cookie  and use urllib3 to get the response,  notice that the faults are in "/doc/html"  which is actually "/data/nginx/html/doc/html/"  directory in APIC (as viewd from root ssh session)
        url = "http://" + apic_ip + "/doc/html/" + fcode
        http = urllib3.PoolManager() #  urllib3.PoolManager doc can be found at:  https://urllib3.readthedocs.io/en/1.2.1/managers.html
        response = http.request('GET', url, headers=cookies)
        # response = urllib3.request('GET', url, headers=cookies)  # could have done this without urllib3.PoolManager
        # *************************************
        #
        # Now use BeautifulSoup4 to get the body and write to file "thisFault.txt"
        soup = bs.BeautifulSoup(response.data, 'lxml')
        fh = open("thisFault.txt", "w")
        body = soup.body
        body_string = body.text.encode('utf-8')  # Noticed some funky encoding, that sometimes screws things up, so converting to utf-8
        fh.write(body_string)
        fh.close()
        time.sleep(1)  #  so many I/O so,  giving it some time
        #
        #   Get the Values of the Fault and make a dictionary of it
        #
        def getValue(key):
        #
            hand = open('thisFault.txt', 'r')  #  open the file for read only
            for line in hand:
                line = line.strip()
                #
                if re.search(key, line):
                    result = line.split(":")
                    mydict[result[0]] = result[1]
                 #
            hand.close()   # close the file
        #
        keys = ['Fault Name:', 'Raised on MO:', 'Weight:', 'Cause:', 'Type:', 'Code:', 'Severity:', "Message:"]
        #
        for key in keys:
            getValue(key)
        #
        #   Some of the fields could not be obtained easily from getValue() function,  so using a different method to populate dictionary
        # ***********************Get Recommended Action: *************************
        #
        str1 = ''  #  creating an empty string to store the value
        hand = open('thisFault.txt', 'r')
        k = 1   # This counter is not really needed,  just in case you wanted to see homw many lines you caught
        #
        #
        lineNum = 0
        for line in hand:
            line = line.strip()
            lineNum += 1
        #
            if re.search("Recommended Action:", line):  # by looking at the file "thisFault.txt" find out where this field is
                # print (line)
                startLine = lineNum
                # print (startLine)
            #
            if re.search("Raised on MO:", line): # by looking at the file "thisFault.txt" find out where this field is
                # print (line)
                stopLine = lineNum - 1
                # print (stopLine)
             #
        hand.seek(0)   #  move the cursor to the 1st position of the file
        #
        lineNum = 0
        for line in hand:
            line = line.strip()
            lineNum += 1
            #
            if lineNum >= startLine and lineNum <= stopLine:
                # print (line)
                str1 = (str1 + line + ' ' + '>' + '.) ') # concenate all the values inbetween the lines obtained
                str1.strip()
                k += 1
                # print ("The number of lines caught is: ", k)
        # print (str1)
        # print (mydict)
        mydict["Recommended Action"] = str1
        #
        #
        print mydict["Recommended Action"]  #  keep this, so you can see that the script is running
        #
        hand.close()
        #
        # *************************************************************************
        #
        # ********************* Get Explanation ***********************
        #
        str1 = '' #  creating an empty string to store the value
        hand = open('thisFault.txt', 'r')
        #
        lineNum = 0
        for line in hand:
            line = line.strip()
            lineNum += 1
            #
            if re.search("Explanation:", line):
                startLine = lineNum
            #
            if re.search("Recommended Action:", line):
                stopLine = lineNum - 1
            #
        hand.seek(0)  #  move the cursor to the 1st position of the file
        #
        lineNum = 0
        for line in hand:
            line = line.strip()
            lineNum += 1
            #
            if lineNum >= startLine and lineNum <= stopLine:
                # print (line)
                str1 = (str1 + line + ' ')
                str1.strip()
        #
        # print (str1)
        #
        mydict["Explanation"] = str1
        # print (mydict)
        # print mydict["Recommended Action"]
        hand.close()
        # *************************************************************
        #
        # ********************* Get Message ***********************
        #
        str1 = '' #  creating an empty string to store the value
        hand = open('thisFault.txt', 'r')
        #
        lineNum = 0
        for line in hand:
            line = line.strip()
            lineNum += 1
            #
            if re.search("Message:", line):
                startLine = lineNum
            #
            if re.search("Help:", line):
                stopLine = lineNum - 1
            #
        hand.seek(0)  #  move the cursor to the 1st position of the file
        #
        #
        lineNum = 0
        for line in hand:
            line = line.strip()
            lineNum += 1
            #
            if lineNum >= startLine and lineNum <= stopLine:
                # print (line)
                str1 = (str1 + line + ' ')
                str1.strip()
             #
        # print (str1)
        #
        mydict["Message"] = str1
        #
        hand.close()
        #
        #  The value of row was obtaine earlier from highest_row
        #
        #  Now populate the spreadsheet
        #
        FirstSheet["A" + str(row)] = mydict['Code']
        FirstSheet["B" + str(row)] = mydict['Message']
        FirstSheet["C" + str(row)] = mydict['Explanation']
        FirstSheet["D" + str(row)] = mydict['Severity']
        FirstSheet["E" + str(row)] = mydict['Weight']
        FirstSheet["F" + str(row)] = mydict['Fault Name']
        FirstSheet["G" + str(row)] = mydict['Cause']
        FirstSheet["H" + str(row)] = mydict['Type']
        FirstSheet["I" + str(row)] = mydict['Raised on MO']
        FirstSheet["j" + str(row)] = mydict['Recommended Action']
        row += 1  # so we can keep writing to the next row for the next fault values
        #
        wb.save("ACI-Faults.xlsx")  # Save the excel spreadsheet after writing each row with fault values
        print ("saving spreadsheet with values of: ", fcode)
        time.sleep(1)  #  give it a break
#
#
#   All Done with writing Faults
#   Auto Align Cells and wrap cells to make the spreadsheet more readable
#
onrow = 1
for row in FirstSheet.iter_rows():
    for cell in row:
        wrap_alignment = Alignment(wrap_text=True)
        cell.alignment = wrap_alignment
    print ("formatting cells on row #  ", onrow)
    onrow += 1
#
wb.save("ACI-Faults.xlsx")  # Save the excel spreadsheet
print "\n" *2
print ("All Done.  Please open the spreadsheet \"ACI-Faults.xlsx\"  now")







