import json
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
from openpyxl.utils import  get_column_letter, column_index_from_string # import column_letter and index from string
#
# fout = open('ACIGroups-csv.txt', 'w')  # open a new file for write


fname = open("ACI-groups.json", "r")
jfile = fname.read()
# print jfile

info = json.loads(jfile)
for headers in info:
    print headers

#
wb = openpyxl.Workbook()  #  open new workbook
#
# *********************consumer-provider  Sheet********************************
# open workbook and create consumer-provider  sheet, Create Header Row
wb.create_sheet(index=0, title='consumer-provider') # insert a new sheet in front
FirstSheet = wb.get_sheet_by_name('consumer-provider')  # choose this sheet
FirstSheet["A1"] = "Consumer"
FirstSheet["B1"] = "Provider"
FirstSheet["C1"] = "Protocol#"
FirstSheet["D1"] = "StartPort#"
FirstSheet["E1"] = "EndPort#"

# Make them bold, and turn AutoFilter on

for i in range(1,6):   #    1 through 5
    columnnum = get_column_letter(i)
    colrow = columnnum + str(1)
    cell = FirstSheet[colrow]
    cell.font = Font(size=12, name="Arial", color=colors.BLUE, italic=True, bold=True)
    FirstSheet.column_dimensions[columnnum].width = 20  #  Set size of Columns to 20
#
# Turn on Freeze Panes and Auto Filter
FirstSheet.freeze_panes = "A2"  # Freeze Panes
FirstSheet.auto_filter.ref = "A1:E1"  # Turn On Auto Filter
#
# Populate  Data in Consumer-Provider Sheet

row = 2
for item in info["policies"]:
            Consumer = item["src_name"]
            Provider = item['dst_name']
            for port in item["whitelist"]:
                Proto = port["proto"]
                Port = port["port"]
                # Write Row 1
                FirstSheet["A" + str(row)]= Consumer
                FirstSheet["B" + str(row)] = Provider
                FirstSheet["C" + str(row)] = str(Proto)
                FirstSheet["D" + str(row)] = str(Port[0])
                FirstSheet["E" + str(row)] = str(Port[1])
                row+=1


# *********************EPG-Grouping  Sheet********************************
#
# Create EPG-Grouping Sheet
wb.create_sheet(index=1, title='EPG-Grouping') # insert a new sheet in front
SecondSheet = wb.get_sheet_by_name('EPG-Grouping')  # choose this sheet
SecondSheet["A1"] = "EPG"
SecondSheet["B1"] = "Provider"
#


# Make them bold, and turn AutoFilter on

for i in range(1,3):   #    1 through 3
    columnnum = get_column_letter(i)
    colrow = columnnum + str(1)
    cell = SecondSheet[colrow]
    cell.font = Font(size=12, name="Arial", color=colors.BLUE, italic=True, bold=True)
    SecondSheet.column_dimensions[columnnum].width = 20  #  Set size of Columns to 20
    #
#  Turn on Freeze Panes and Auto Filter
SecondSheet.freeze_panes = "A2"  # Freeze Panes
SecondSheet.auto_filter.ref = "A1:B1"  # Turn On Auto Filter
#

# Populate  Data in EPG-Grouping Sheet
row = 2
for item in info["clusters"]:
          EPG = item["name"]
          for node in item["nodes"]:
            IPADD = node["ip"]
            # STR1 = EPG + "~" + IPADD
            SecondSheet["A" + str(row)] = EPG
            SecondSheet["B" + str(row)] = IPADD
            row+=1

# **********************************************
#Remove the last sheet,  name of Sheet
wb.remove_sheet(wb.get_sheet_by_name('Sheet'))


wb.save("tetration_json_output2excel.xlsx")  # Save the excel spreadsheet

